from subprocess import CalledProcessError, run as SubprocessRun
from openpyxl import Workbook, load_workbook
#from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from re import sub, search
from tkinter import filedialog
from os import listdir, path as OsPath
import zipfile
from numpy import where, loadtxt
from datetime import datetime



#######################################################################################################
#BEGIN FUNCTIONS

def AdjustABIChars(fileName):
    newFileName = fileName
    newFileName = newFileName.replace(' ', '')
    newFileName = newFileName.replace('+', '&')
    newFileName = newFileName.replace("*", "-")
    newFileName = newFileName.replace("|", "-")
    newFileName = newFileName.replace("/", "-")
    newFileName = newFileName.replace("\\", "-")
    newFileName = newFileName.replace(":", "-")
    newFileName = newFileName.replace("\"", "")
    newFileName = newFileName.replace("\'", "")
    newFileName = newFileName.replace("<", "-")
    newFileName = newFileName.replace(">", "-")
    newFileName = newFileName.replace("?", "")
    newFileName = newFileName.replace(",", "")
    return newFileName

#Need to compare sample_primer names to names adjusted for our "Premixed" and "RTI" endings.
def NeutralizeSuffixesAnd_ab1_Extension(fileName):
    newFileName = fileName
    newFileName = newFileName.replace('_Premixed', '')
    newFileName = newFileName.replace('_RTI', '')
    newFileName = newFileName.replace('.ab1', '')
    return newFileName

def GetInumberFolders(path):
    INumberPaths = []
    for item in listdir(path):                              #for all folders in the user-selected daily data folder                                           
        if OsPath.isdir(f'{path}\\{item}'):                    #if item is a folder
            '''print(search('bioi-\\d+', item.lower()))
            print(item.lower()[:4])
            print(search('reinject', item.lower()))
            print(search('bioi-\\d+_.+_\\d+', item.lower()))'''
            if search('bioi-\\d+', item.lower()) and (item.lower()[:4]=='bioi') and (search('reinject', item.lower())==None) and (search('bioi-\\d+_.+_\\d+', item.lower())==None):       
            #if item looks like a BioI- folder. Only want to get bioI folders that have been sorted - looking like "BioI-20000" and don't include order folders here. We will handle order folders that are in the main folder separately.
                INumberPaths.append(f'{path}\\{item}')
    return  INumberPaths

#returns a list of paths to order folders in the given path
def GetImmediateOrders(path):
    orderPaths = []
    for item in listdir(path):
        if OsPath.isdir(f'{path}\\{item}'):
            if search('bioi-\\d+_.+_\\d+', item.lower()) and not 'reinject' in item.lower():
                orderPaths.append(f'{path}\\{item}')
    return orderPaths

#function gets full paths of the order folders given a BioI folder. 
# returns list of full paths of order folders. 
def GetOrderFolders(plateFolder):
    orderFolderPaths = []

    #I don't think this would capture any by the way the main loop works. Also we'll handle immediate orders(not in order folders) 
    #if search('bioi-\\d+_.+_\\d', plateFolder.lower()): #if "plateFolder" itself is an order
    #    orderFolderPaths.append(f'{plateFolder}')

    for item in listdir(plateFolder):
        if OsPath.isdir(f'{plateFolder}\\{item}'):
            if search('bioi-\\d+_.+_\\d+', item.lower()) and (search('reinject', item.lower())==None):            #if item looks like an order folder BioI-20000_YoMama_123456 Also don't get any that say reinject.
                orderFolderPaths.append(f'{plateFolder}\\{item}')   #add to list
    return orderFolderPaths

def GetZip(path):
    zipfilepath = ''
    for item in listdir(path):
        if item.endswith('.zip'):
            zipfilepath = f'{path}\\{item}'
            break
    return zipfilepath

def GetOrderList(number):
    orderSampleNames = []
    orderIndexes = where((key == str(number)))
    i = 0 
    while i < len(orderIndexes[0]):
        rawName = key[orderIndexes[0][i], 3]
        adjustedName = AdjustABIChars(rawName)
        orderSampleNames.append([adjustedName, rawName])
        i+=1
    return orderSampleNames

#function finds and returns just the order number of the BioI folder name
def GetOrderNumberFromFolderName(folder):
    orderNum = ''
    if search('.+_\\d+', folder):
        orderNum = search('_\\d+', folder).group(0)
        orderNum = search('\\d+', orderNum).group(0)
        return orderNum
    else:
        return ''
    
def GetZipModTime(sheet, number):
    foundModTime = ''
    for rowData in sheet.iter_rows(values_only=True):
        if str(rowData[1])==str(number):
            foundModTime = rowData[7]
            break
    return foundModTime

def GetExistingRowsHidden(sheet):
    rowshidden = []
    i = 2 
    existTotalRows = sheet.max_row
    while i < existTotalRows: rowshidden.append(sheet.row_dimensions[i].hidden); i+=1
    return rowshidden
    

def ResolveOrder(sheet, number):
    #status = "resolved"
    #status color = resolvedStyle
    #hid the rows.
    startHidingRows = False
    boolfornext = False
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if str(row[1])==str(number):
            orow = i
            #row[2] = "Resolved"
            #row[2].fill = resolvedStyle
            sheet[f'C{i}'] = "Resolved"
            sheet[f'C{i}'].fill = resolvedStyle
            boolfornext = True
        if boolfornext and i == orow + 1: #start hiding rows 1 row after the order row
            startHidingRows = True
        if startHidingRows and sheet[f'B{i}'].value==None:
            sheet.row_dimensions[i].hidden = True #we'll hide rows until we hit another order in the sheet, or by the nature of the loop, until we hit the last row
        else:
            if startHidingRows:
                break
     

#END FUNCTIONS
######################################################################################################


if __name__ == "__main__":
    batPath = "P:\\generate-data-sorting-key-file.bat"
    keyFilePathName = 'P:\\order_key.txt'
    successStyle = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
    attentionStyle = PatternFill(start_color='FF4747', end_color='FF4747', fill_type='solid')
    resolvedStyle = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    breakStyle = PatternFill(start_color='DDD9C4', end_color='DDD9C4', fill_type='solid')

    rowCount = 1
    txtlist = ['raw.qual.txt', 'raw.seq.txt', 'seq.info.txt', 'seq.qual.txt', 'seq.txt']
    timeStamp = str(int(datetime.now().timestamp()))[-4:]
    sumExists = False

    try:
        SubprocessRun(batPath, shell=True, check=True)
        dataFolderPath = filedialog.askdirectory(title="Select today's data folder to validate zip files.")
        dataFolderPath = sub(r'/','\\\\', dataFolderPath)
        key = loadtxt(keyFilePathName, dtype=str, delimiter='\t')

        bioIFolders  = GetInumberFolders(dataFolderPath)  #get bioIfolders that have been sorted already
        immediateOrderFolders = GetImmediateOrders(dataFolderPath)
        
        savePath = dataFolderPath
        bookName = f'zip order summary - {datetime.now().strftime("%Y-%m-%d")}.xlsx'
        sumExists = OsPath.exists(f'{savePath}\\{bookName}')

        #we'll add all new data to this book regardless of summary existing. 
        summaryBook = Workbook()
        sumSheet = summaryBook.active 

        if sumExists:
            existingBook = load_workbook(f'{savePath}\\{bookName}')
            existSheet = existingBook.active
            
            #store the hidden status of the rows and unhide all of them, so they have a fresh reset. 
            existingRowshidden = GetExistingRowsHidden(existSheet)
        else: 
            #put headers in 
            sumSheet[f'A{1}'] = 'I Number'
            sumSheet[f'B{1}'] = 'Order Number'
            sumSheet[f'E{1}'] = 'Order Items'
            sumSheet[f'F{1}'] = 'File Names'
            sumSheet[f'G{1}'] = 'Status'
            sumSheet[f'H{1}'] = 'Zip Timestamp'
            rowCount +=1
        


        for bioIFolder in bioIFolders:
            orderFolders = GetOrderFolders(bioIFolder) #get order folders which will be compressed.
            for orderFolder in orderFolders:
                matchCounter = 0
                mismatchCounter = 0
                txtCounter = 0
                orderIsInSummary = False
                zipModTime = ''

                #get order information from key
                orderNumber = GetOrderNumberFromFolderName(orderFolder)
                iNumber = OsPath.basename(bioIFolder)
                orderKey = GetOrderList(orderNumber) #this contains both ABI adjusted names and the raw names. 
                orderLength = len(orderKey)
                zipPath = GetZip(orderFolder)
                if len(zipPath) > 0: currentModTime = OsPath.getmtime(zipPath)


                if sumExists: 
                    zipModTime = GetZipModTime(existSheet, orderNumber)
                    if len(str(zipModTime)) > 0:
                        orderIsInSummary = True 

                if len(zipPath) > 0:
                    if not sumExists or (orderIsInSummary and (int(zipModTime) < int(currentModTime))) or (sumExists and not orderIsInSummary):
                        #get contents of zip files
                        with zipfile.ZipFile(zipPath, 'r') as zip_ref:
                            zipContents = zip_ref.namelist()
                        #print(zipContents)
                        sumSheet[f'A{rowCount}'] = iNumber
                        sumSheet[f'B{rowCount}'] = orderNumber
                        #C is reserved for the status later. 
                        sumSheet[f'D{rowCount}'] = OsPath.basename(zipPath)
                        orderRow = rowCount
                        rowCount +=1

                        #this loop will list out all the matches first, and store all the unnaccounted items in lists 
                        for item in orderKey[:]:
                            #print(item)
                            #print(item[0]) #abi ajusted
                            #print(item[1]) #raw

                            for zipitem in zipContents[:]:
                                #only address ab1
                                if zipitem.endswith('.ab1'):
                                    # need to compare zipcontents(-.abi, and -RTI -Premixed) to abi the abi adjusted orderKey names
                                    cleanZipItem = NeutralizeSuffixesAnd_ab1_Extension(zipitem)
                                    if item[0]==cleanZipItem:  #match
                                        sumSheet[f'E{rowCount}'] = item[1]  #raw
                                        sumSheet[f'F{rowCount}'] = zipitem
                                        sumSheet[f'G{rowCount}'] = 'match'
                                        sumSheet[f'G{rowCount}'].fill = successStyle
                                        matchCounter += 1
                                        rowCount +=1
                                        #if we have a match, remove
                                        zipContents.remove(zipitem)
                                        orderKey.remove(item)
                                        break

                        #output any remaining items from zipContents
                        for zipitem in zipContents:
                            if zipitem.endswith('.ab1'):
                                sumSheet[f'F{rowCount}'] = zipitem
                                sumSheet[f'G{rowCount}'] = 'no match'
                                sumSheet[f'G{rowCount}'].fill = attentionStyle
                                rowCount +=1
                                mismatchCounter +=1

                        #output any remaining items from order key
                        for item in orderKey:
                            sumSheet[f'E{rowCount}'] = item[1]  #raw
                            sumSheet[f'G{rowCount}'] = 'no match'
                            sumSheet[f'G{rowCount}'].fill = attentionStyle
                            rowCount +=1
                            mismatchCounter +=1

                        #list the txt files at the end. 
                        for txt in txtlist:
                            sumSheet[f'E{rowCount}'] = txt  #txt ending
                            for zipitem in zipContents:
                                if zipitem.endswith(txt):
                                    sumSheet[f'F{rowCount}'] = zipitem
                                    zipContents.remove(zipitem)
                                    sumSheet[f'G{rowCount}'] = 'txt file'
                                    sumSheet[f'G{rowCount}'].fill = successStyle
                                    txtCounter +=1
                                    break
                            else:
                                sumSheet[f'G{rowCount}'] = 'MISSING txt file'
                                sumSheet[f'G{rowCount}'].fill = attentionStyle
                            rowCount +=1
                        if ('andreev' in orderFolder.lower())==False:
                            if (orderLength == matchCounter) and txtCounter==5 and matchCounter != 0:
                                #if was found to be in existing sheet and modzip < current zip (and we are here in completed territory) then we can change the status in the existing sheet
                                if orderIsInSummary and (int(zipModTime) < int(currentModTime)):
                                    ResolveOrder(existSheet, orderNumber)
                                    #update rows hidden
                                    existingRowshidden = GetExistingRowsHidden(existSheet)

                                sumSheet[f'C{orderRow}'] = 'Completed'
                                sumSheet[f'C{orderRow}'].fill = successStyle
                                sumSheet[f'H{orderRow}'] = currentModTime
                                i = orderRow + 1
                                while i < rowCount:
                                    sumSheet.row_dimensions[i].hidden = True
                                    i+=1
                            else:
                                sumSheet[f'C{orderRow}'] = 'ATTENTION'
                                sumSheet[f'C{orderRow}'].fill = attentionStyle
                                sumSheet[f'H{orderRow}'] = currentModTime
                        else: #disregard the txt files for Dmitri
                            if (orderLength == matchCounter) and matchCounter != 0: 
                                if orderIsInSummary and (int(zipModTime) < int(currentModTime)):
                                    ResolveOrder(existSheet, orderNumber)
                                    #update rows hidden
                                    existingRowshidden = GetExistingRowsHidden(existSheet)
                                sumSheet[f'C{orderRow}'] = 'Completed'
                                sumSheet[f'C{orderRow}'].fill = successStyle
                                sumSheet[f'H{orderRow}'] = currentModTime
                                i = orderRow + 1
                                while i < rowCount:
                                    sumSheet.row_dimensions[i].hidden = True
                                    i+=1
                            else:
                                sumSheet[f'C{orderRow}'] = 'ATTENTION'
                                sumSheet[f'C{orderRow}'].fill = attentionStyle
                                sumSheet[f'H{orderRow}'] = currentModTime

        #fit column widths. It'
        if rowCount > 2: 
            #savePath = 'H:\\'
            #bookName = 'summary.xlsx'
            for col in sumSheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col: 
                    try: 
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except: 
                        pass
                adjusted_width = max_length + 2
                sumSheet.column_dimensions[column].width = adjusted_width
            if  sumExists:
                sumSheet[f'A{rowCount}'] = "Break"
                sumSheet[f'A{rowCount}'].fill = breakStyle
                sumSheet.row_dimensions[rowCount].hidden = False
    
        savePath = dataFolderPath
        bookName = f'zip order summary - {datetime.now().strftime("%Y-%m-%d")}.xlsx'
        if not sumExists:
            summaryBook.save(f'{savePath}\\{bookName}')
        else: #we're gonna use the same code below but we'll make the data in a new virtual workbook, then insert that data into the existing workbook
            #get the data from our summary(including the colors), and put it into a list.
            additionalData = []
            for i, row in enumerate(sumSheet.iter_rows(values_only=False), start=1):
                rowData = []
                hidden = sumSheet.row_dimensions[i].hidden
                for cell in row:
                    cellfill = cell.fill
                    cellData = {
                        'value': cell.value,
                        'fill': {'patType': cellfill.fill_type, 'scolor': cellfill.start_color, 'ecolor': cellfill.end_color}
                    }
                    rowData.append(cellData)
                additionalData.append([rowData, hidden])

            numRows = len(additionalData)
            existingBook.save(f'{savePath}\\{bookName}')
            existSheet.insert_rows(2, numRows)                                             #insert the rows we need. 2 extra to give a divide between new and old zip data.
            existingBook.save(f'{savePath}\\{bookName}')
            for i, rowData in enumerate(additionalData, start=2):                           #loop to put the data in the existing sheet
                existSheet.row_dimensions[i].hidden = rowData[1]
                for colNum, cellData in enumerate(rowData[0], start=1):
                    cell = existSheet.cell(row=i, column=colNum, value=cellData['value'], ) #values
                    cell.fill = PatternFill(start_color=cellData['fill']['scolor'], end_color=cellData['fill']['ecolor'], patternType=cellData['fill']['patType'])
                #existingBook.save(f'{savePath}\\{bookName}')
            for colLetter, colDimension in sumSheet.column_dimensions.items():               #adjust the target summary sheet column widths to the new sheet widths. 
                existSheet.column_dimensions[colLetter].width = colDimension.width
            for i, rowItem in enumerate(existingRowshidden, start=2):
                existSheet.row_dimensions[i+numRows].hidden = rowItem
            existingBook.save(f'{savePath}\\{bookName}')


    except CalledProcessError:
        print(f'Batch file, {batPath} didn\'t work.')
    except FileNotFoundError:
        print('File not found')
    except PermissionError:
        print('Please close the summary workbook and try again.')

    