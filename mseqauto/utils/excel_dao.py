# excel_dao.py - Simplified version
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelDAO:
    def __init__(self, config):
        self.config = config
        # Define standard styles - keeping exact same colors as original
        self.success_style = PatternFill(start_color='00CC00', end_color='00CC00', fill_type='solid')
        self.attention_style = PatternFill(start_color='FF4747', end_color='FF4747', fill_type='solid')
        self.resolved_style = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        self.break_style = PatternFill(start_color='DDD9C4', end_color='DDD9C4', fill_type='solid')
        
        # Store hidden row states for preservation during updates
        self._stored_hidden_states = {}

    # Core workbook operations
    def create_workbook(self):
        """Create a new workbook"""
        return Workbook()

    def load_workbook(self, file_path):
        """Load an existing workbook with error handling"""
        if not Path(file_path).exists():
            return None
        try:
            return load_workbook(file_path)
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return None

    def save_with_error_handling(self, workbook, file_path):
        """Save workbook with specific error handling for permission issues"""
        try:
            workbook.save(file_path)
            return True
        except PermissionError:
            print(f"Error: Cannot save Excel file. Please close {Path(file_path).name} in Excel and try again.")
            return False
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False

    # Cell operations
    def set_cell_value(self, worksheet, row, col, value):
        """Set cell value"""
        worksheet.cell(row=row, column=col, value=value)

    def get_cell_value(self, worksheet, row, col):
        """Get cell value"""
        return worksheet.cell(row=row, column=col).value

    def apply_style(self, worksheet, cell_ref, style_type):
        """Apply style to cell"""
        style_map = {
            'success': self.success_style,
            'attention': self.attention_style,
            'break': self.break_style,
            'resolved': self.resolved_style
        }
        if style_type in style_map:
            worksheet[cell_ref].fill = style_map[style_type]

    # Formatting operations
    def set_validation_headers(self, worksheet):
        """Set headers for validation summary"""
        headers = ['I Number', 'Order Number', 'Status', 'Zip Filename',
                   'Order Items', 'File Names', 'Match Status', 'Zip Timestamp']
        for i, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    def set_fb_pcr_headers(self, worksheet):
        """Set headers for FB-PCR summary"""
        headers = ['PCR Number', 'Order Number', 'Version', 'Zip Filename', 
                   'Total Files', 'File Types', 'Zip Timestamp']
        for i, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    def set_plate_headers(self, worksheet):
        """Set headers for plate folder summary"""
        headers = ['Plate Number', 'Description', 'Zip Filename', 'Total Files', 
                   'AB1 Files', 'FSA Files', 'File Types', 'Zip Timestamp']
        for i, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    def auto_adjust_columns(self, worksheet):
        """Adjust column widths based on content - matches original script logic"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Row management for updates
    def preserve_hidden_row_states(self, worksheet):
        """Store the current hidden state of all rows"""
        self._stored_hidden_states = {}
        for i in range(2, worksheet.max_row + 1):  # Start from 2 to skip header
            self._stored_hidden_states[i] = worksheet.row_dimensions[i].hidden

    def restore_hidden_row_states(self, worksheet, offset=0):
        """Restore hidden states with optional offset for inserted rows"""
        for original_row, hidden_state in self._stored_hidden_states.items():
            new_row = original_row + offset
            if new_row <= worksheet.max_row:
                worksheet.row_dimensions[new_row].hidden = hidden_state

    def insert_rows_at_top(self, worksheet, num_rows):
        """Insert rows at top (after header) and handle hidden state preservation"""
        if num_rows <= 0:
            return
        self.preserve_hidden_row_states(worksheet)
        worksheet.insert_rows(2, num_rows)

    # Order management
    def find_order_in_summary(self, worksheet, order_number):
        """Find if an order already exists in the summary"""
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[1]) == str(order_number):  # Column B is order number
                zip_timestamp = row[7] if len(row) > 7 else None  # Column H is zip timestamp
                return True, row_num, zip_timestamp
        return False, None, None

    def resolve_order_status(self, worksheet, order_number):
        """Mark an order as resolved and hide associated rows"""
        found, order_row, _ = self.find_order_in_summary(worksheet, order_number)
        
        if not found:
            return False
            
        # Set status to "Resolved"
        self.set_cell_value(worksheet, order_row, 3, "Resolved")  # Column C
        self.apply_style(worksheet, f'C{order_row}', 'resolved')
        
        # Hide rows until we hit another order or end of data
        current_row = order_row + 1 #type: ignore
        while current_row <= worksheet.max_row:
            order_cell_value = self.get_cell_value(worksheet, current_row, 2)
            if order_cell_value is not None and str(order_cell_value).strip():
                break  # Hit another order, stop hiding
            else:
                worksheet.row_dimensions[current_row].hidden = True
                current_row += 1
                
        return True

    def add_break_row(self, worksheet, row_num):
        """Add a break row with styling"""
        self.set_cell_value(worksheet, row_num, 1, "Break")
        self.apply_style(worksheet, f'A{row_num}', 'break')
        worksheet.row_dimensions[row_num].hidden = False

    # Main validation result processing
    def add_validation_result(self, worksheet, row_count, validation_result, zip_path, i_number, order_number,
                              is_andreev=False):
        """Add validation result to worksheet"""
        # Set basic information
        self.set_cell_value(worksheet, row_count, 1, i_number)
        self.set_cell_value(worksheet, row_count, 2, order_number)
        self.set_cell_value(worksheet, row_count, 4, Path(zip_path).name)
        self.set_cell_value(worksheet, row_count, 8, str(int(Path(zip_path).stat().st_mtime)))

        order_row = row_count
        row_count += 1

        # Determine status
        match_count = validation_result.get('match_count', 0)
        expected_count = validation_result.get('expected_count', 0)
        txt_count = validation_result.get('txt_count', 0)
        extra_ab1_count = validation_result.get('extra_ab1_count', 0)

        # Status logic: completed if all expected files match, no extra files, and (for non-Andreev) all txt files present
        is_completed = (match_count == expected_count and match_count != 0 and extra_ab1_count == 0 and
                       (is_andreev or txt_count == 5))

        status = 'Completed' if is_completed else 'ATTENTION'
        style = 'success' if is_completed else 'attention'
        
        self.set_cell_value(worksheet, order_row, 3, status)
        self.apply_style(worksheet, f'C{order_row}', style)

        # Add file details
        for match in validation_result.get('matches', []):
            self.set_cell_value(worksheet, row_count, 5, match['raw_name'])
            self.set_cell_value(worksheet, row_count, 6, match['file_name'])
            self.set_cell_value(worksheet, row_count, 7, 'match')
            self.apply_style(worksheet, f'G{row_count}', 'success')
            row_count += 1

        for mismatch in validation_result.get('mismatches_in_zip', []):
            self.set_cell_value(worksheet, row_count, 6, mismatch)
            self.set_cell_value(worksheet, row_count, 7, 'no match')
            self.apply_style(worksheet, f'G{row_count}', 'attention')
            row_count += 1

        for mismatch in validation_result.get('mismatches_in_order', []):
            self.set_cell_value(worksheet, row_count, 5, mismatch['raw_name'])
            self.set_cell_value(worksheet, row_count, 7, 'no match')
            self.apply_style(worksheet, f'G{row_count}', 'attention')
            row_count += 1

        # Add txt file status
        for txt_ext in self.config.TEXT_FILES:
            self.set_cell_value(worksheet, row_count, 5, txt_ext)
            if txt_ext in validation_result.get('txt_files', []):
                self.set_cell_value(worksheet, row_count, 6, f"*{txt_ext}")
                self.set_cell_value(worksheet, row_count, 7, 'txt file')
                self.apply_style(worksheet, f'G{row_count}', 'success')
            else:
                self.set_cell_value(worksheet, row_count, 7, 'MISSING txt file')
                self.apply_style(worksheet, f'G{row_count}', 'attention')
            row_count += 1

        # Hide rows for completed orders
        if is_completed:
            for i in range(order_row + 1, row_count):
                worksheet.row_dimensions[i].hidden = True

        return row_count

    def add_fb_pcr_result(self, worksheet, row_count, fb_pcr_result, zip_path, mixed_headers=False):
        """Add FB-PCR result to worksheet with individual file listing"""
        order_row = row_count  # Store the main row for hiding purposes
        
        if mixed_headers:
            # Use validation headers format for mixed data
            # I Number, Order Number, Status, Zip Filename, Order Items, File Names, Match Status, Zip Timestamp
            self.set_cell_value(worksheet, row_count, 1, f"PCR-{fb_pcr_result['pcr_number']}")  # I Number -> PCR Number
            self.set_cell_value(worksheet, row_count, 2, fb_pcr_result['order_number'])        # Order Number
            self.set_cell_value(worksheet, row_count, 3, "FB-PCR")                            # Status -> FB-PCR
            self.set_cell_value(worksheet, row_count, 4, Path(zip_path).name)                 # Zip Filename
            
            # Show .ab1 file count prominently, with text file count as additional info
            ab1_count = fb_pcr_result.get('ab1_count', 0)
            txt_count = fb_pcr_result.get('txt_count', 0)
            total_files = fb_pcr_result.get('total_files', 0)
            self.set_cell_value(worksheet, row_count, 5, f".ab1 Files: {ab1_count}, Text Files: {txt_count}, Total: {total_files}")  # Order Items -> File counts
            
            # Format file types as a readable string for File Names column
            file_types_str = ", ".join([f"{ext}: {count}" for ext, count in fb_pcr_result['file_types'].items()])
            self.set_cell_value(worksheet, row_count, 6, file_types_str)                      # File Names -> File types
            
            self.set_cell_value(worksheet, row_count, 7, f"Version {fb_pcr_result['version']}")  # Match Status -> Version
            self.set_cell_value(worksheet, row_count, 8, str(int(Path(zip_path).stat().st_mtime)))  # Zip Timestamp
            
            # Apply styling for FB-PCR entries
            self.apply_style(worksheet, f'C{row_count}', 'success')
        else:
            # Use FB-PCR specific headers format
            # PCR Number, Order Number, Version, Zip Filename, Total Files, File Types, Zip Timestamp
            self.set_cell_value(worksheet, row_count, 1, fb_pcr_result['pcr_number'])
            self.set_cell_value(worksheet, row_count, 2, fb_pcr_result['order_number'])
            self.set_cell_value(worksheet, row_count, 3, fb_pcr_result['version'])
            self.set_cell_value(worksheet, row_count, 4, Path(zip_path).name)
            
            # Show .ab1 file count prominently in the Total Files column
            ab1_count = fb_pcr_result.get('ab1_count', 0)
            txt_count = fb_pcr_result.get('txt_count', 0)
            total_files = fb_pcr_result.get('total_files', 0)
            self.set_cell_value(worksheet, row_count, 5, f".ab1: {ab1_count}, Text: {txt_count}, Total: {total_files}")
            
            # Format file types as a readable string
            file_types_str = ", ".join([f"{ext}: {count}" for ext, count in fb_pcr_result['file_types'].items()])
            self.set_cell_value(worksheet, row_count, 6, file_types_str)
            
            self.set_cell_value(worksheet, row_count, 7, str(int(Path(zip_path).stat().st_mtime)))

        row_count += 1

        # Add individual file names for verification (similar to validation results)
        # Sort files so .ab1 files appear first, then text files
        file_names = fb_pcr_result.get('file_names', [])
        ab1_files = [f for f in file_names if f.lower().endswith('.ab1')]
        text_files = [f for f in file_names if not f.lower().endswith('.ab1')]
        sorted_files = ab1_files + text_files
        
        for file_name in sorted_files:
            if mixed_headers:
                # Use columns 5 and 6 for file details in mixed format
                self.set_cell_value(worksheet, row_count, 5, "")  # Empty order items column
                self.set_cell_value(worksheet, row_count, 6, file_name)  # File name in file names column
            else:
                # Use available columns for FB-PCR format - put in file types column for now
                self.set_cell_value(worksheet, row_count, 6, file_name)
            row_count += 1

        # Hide the individual file name rows (keeping only the summary row visible)
        for i in range(order_row + 1, row_count):
            worksheet.row_dimensions[i].hidden = True

        return row_count

    def add_plate_result(self, worksheet, row_count, plate_result, zip_path, mixed_headers=False):
        """Add plate folder result to worksheet with individual file listing"""
        order_row = row_count  # Store the main row for hiding purposes
        
        if mixed_headers:
            # Use validation headers format for mixed data
            # I Number, Order Number, Status, Zip Filename, Order Items, File Names, Match Status, Zip Timestamp
            self.set_cell_value(worksheet, row_count, 1, f"P{plate_result['plate_number']}")  # I Number -> Plate Number
            self.set_cell_value(worksheet, row_count, 2, plate_result['description'])         # Order Number -> Description
            self.set_cell_value(worksheet, row_count, 3, "PLATE")                            # Status -> PLATE
            self.set_cell_value(worksheet, row_count, 4, Path(zip_path).name)                # Zip Filename
            
            # Show file counts prominently
            ab1_count = plate_result.get('ab1_count', 0)
            fsa_count = plate_result.get('fsa_count', 0)
            txt_count = plate_result.get('txt_count', 0)
            total_files = plate_result.get('total_files', 0)
            self.set_cell_value(worksheet, row_count, 5, f".ab1: {ab1_count}, .fsa: {fsa_count}, Text: {txt_count}, Total: {total_files}")  # Order Items -> File counts
            
            # Format file types as a readable string for File Names column
            file_types_str = ", ".join([f"{ext}: {count}" for ext, count in plate_result['file_types'].items()])
            self.set_cell_value(worksheet, row_count, 6, file_types_str)                      # File Names -> File types
            
            self.set_cell_value(worksheet, row_count, 7, "PLATE")                             # Match Status -> PLATE
            self.set_cell_value(worksheet, row_count, 8, str(int(Path(zip_path).stat().st_mtime)))  # Zip Timestamp
            
            # Apply styling for plate entries
            self.apply_style(worksheet, f'C{row_count}', 'success')
        else:
            # Use plate specific headers format
            # Plate Number, Description, Zip Filename, Total Files, AB1 Files, FSA Files, File Types, Zip Timestamp
            self.set_cell_value(worksheet, row_count, 1, plate_result['plate_number'])
            self.set_cell_value(worksheet, row_count, 2, plate_result['description'])
            self.set_cell_value(worksheet, row_count, 3, Path(zip_path).name)
            
            # Show file counts in Total Files column
            ab1_count = plate_result.get('ab1_count', 0)
            fsa_count = plate_result.get('fsa_count', 0)
            txt_count = plate_result.get('txt_count', 0)
            total_files = plate_result.get('total_files', 0)
            self.set_cell_value(worksheet, row_count, 4, f"Total: {total_files}")
            self.set_cell_value(worksheet, row_count, 5, f".ab1: {ab1_count}")
            self.set_cell_value(worksheet, row_count, 6, f".fsa: {fsa_count}")
            
            # Format file types as a readable string
            file_types_str = ", ".join([f"{ext}: {count}" for ext, count in plate_result['file_types'].items()])
            self.set_cell_value(worksheet, row_count, 7, file_types_str)
            
            self.set_cell_value(worksheet, row_count, 8, str(int(Path(zip_path).stat().st_mtime)))

        row_count += 1

        # Add individual file names for verification (similar to validation results)
        # Sort files so .ab1 files appear first, then .fsa files, then text files
        file_names = plate_result.get('file_names', [])
        ab1_files = [f for f in file_names if f.lower().endswith('.ab1')]
        fsa_files = [f for f in file_names if f.lower().endswith('.fsa')]
        text_files = [f for f in file_names if f.lower().endswith('.txt')]
        other_files = [f for f in file_names if not any(f.lower().endswith(ext) for ext in ['.ab1', '.fsa', '.txt'])]
        
        sorted_files = ab1_files + fsa_files + text_files + other_files
        
        for file_name in sorted_files:
            if mixed_headers:
                # Use columns 5 and 6 for file details in mixed format
                self.set_cell_value(worksheet, row_count, 5, "")  # Empty order items column
                self.set_cell_value(worksheet, row_count, 6, file_name)  # File name in file names column
            else:
                # Use available columns for plate format - put in file types column for now
                self.set_cell_value(worksheet, row_count, 7, file_name)
            row_count += 1

        # Hide the individual file name rows (keeping only the summary row visible)
        for i in range(order_row + 1, row_count):
            worksheet.row_dimensions[i].hidden = True

        return row_count

    # Complex update operations
    def copy_data_with_formatting(self, source_worksheet, start_row=2):
        """Copy data with formatting from source worksheet"""
        copied_rows = []
        
        for i, row in enumerate(source_worksheet.iter_rows(min_row=start_row, values_only=False), start=start_row):
            row_data = []
            hidden = source_worksheet.row_dimensions[i].hidden
            
            for cell in row:
                cell_fill = cell.fill
                cell_data = {
                    'value': cell.value,
                    'fill': {
                        'fill_type': cell_fill.fill_type,
                        'start_color': cell_fill.start_color,
                        'end_color': cell_fill.end_color
                    }
                }
                row_data.append(cell_data)
            
            copied_rows.append([row_data, hidden])
        
        return copied_rows

    def paste_data_with_formatting(self, worksheet, data_rows, start_row=2):
        """Paste data with formatting to worksheet"""
        for i, (row_data, hidden) in enumerate(data_rows, start=start_row):
            worksheet.row_dimensions[i].hidden = hidden
            
            for col_num, cell_data in enumerate(row_data, start=1):
                cell = worksheet.cell(row=i, column=col_num, value=cell_data['value'])
                cell.fill = PatternFill(
                    start_color=cell_data['fill']['start_color'],
                    end_color=cell_data['fill']['end_color'],
                    fill_type=cell_data['fill']['fill_type']
                )

    def update_existing_summary(self, existing_workbook, new_data_workbook, save_path):
        """Update existing summary with new data"""
        existing_sheet = existing_workbook.active
        new_sheet = new_data_workbook.active
        
        # Get new data with formatting
        new_data_rows = self.copy_data_with_formatting(new_sheet)
        num_new_rows = len(new_data_rows)
        
        if num_new_rows == 0:
            return True
            
        # Insert rows and handle data transfer
        self.insert_rows_at_top(existing_sheet, num_new_rows)
        self.paste_data_with_formatting(existing_sheet, new_data_rows, start_row=2)
        
        # Copy column widths
        for col_letter, col_dimension in new_sheet.column_dimensions.items():
            existing_sheet.column_dimensions[col_letter].width = col_dimension.width
        
        # Restore hidden states with offset
        self.restore_hidden_row_states(existing_sheet, offset=num_new_rows)
        
        return self.save_with_error_handling(existing_workbook, save_path)

    def finalize_workbook(self, worksheet, add_break_at_end=False):
        """Finalize workbook with column adjustments and optional break row"""
        self.auto_adjust_columns(worksheet)
        if add_break_at_end:
            self.add_break_row(worksheet, worksheet.max_row + 1)


if __name__ == "__main__":
    # Simplified test
    class TestConfig:
        TEXT_FILES = ['.raw.qual.txt', '.raw.seq.txt', '.seq.info.txt', '.seq.qual.txt', '.seq.txt']

    config = TestConfig()
    excel_dao = ExcelDAO(config)
    
    wb = excel_dao.create_workbook()
    ws = wb.active
    excel_dao.set_validation_headers(ws)
    
    print("Simplified ExcelDAO test completed!")